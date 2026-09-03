#!/usr/bin/env python3
"""
Builds index.html for the SIORE 재고 운영 대시보드 from the live Google Sheet.
Run by .github/workflows/update.yml on a schedule and via the sheet-updated webhook.
"""
import os
import sys
import json
import calendar as calmod
import datetime
from collections import OrderedDict

import requests
import openpyxl

SHEET_ID = os.environ.get("GOOGLE_SHEET_ID", "1Tm-hxHQ_Uzw_Hn7wl5Yr0iyh5W80VuwKTISHryTveYU")
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
XLSX_PATH = "/tmp/_source.xlsx"
OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "index.html")

# SIORE logo, base64 PNG (white wordmark) - kept inline so the page has no external deps.
LOGO_DATAURI = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAd4AAACKCAYAAAAE9iuWAAAIeUlEQVR4nO3dzZajSA4G0PCcfv9X9iyyOe10pSuNiR9Jce9mNtOVgIQ+Aht8u9/vDQCY43+rNwAAdiJ4AWAiwQsAEwleAJhI8ALARIIXACYSvAAw0T+rNwAAJuj10orb1X/AiheA6sKEbmuCF4DaQoVua241A1BTuMA9WPECUE3Y0G1N8AJQS+jQbU3wAlBH+NBtzWe8ANTQI3SHBu7BiheA7NKEbmuCF4DcUoVua241A5DX1dCdGrgHK14AMkoZuq0JXgDySRu6rQleAHJJHbqt+YwXgDyuhO7ywD1Y8QKQQYnQbU3wAhBfmdBtza1mAGL7NHTDBe7BiheAasKGbmtWvADEFjpEP2HFCwATCV4AmEjwAsBEghcAJsr25arfvlZe7kN4QtKHwMciBW+P31TcdSCWe85tsau9uGsfXtVjBoyiZq9FrtsKv/bKyuBdUayf/qYTitm9qA/zOdMjaslfzQ7eiFdGj9vkhNlHtF7Uh3U895Z68s2s4I025F4x/OrL0IvHNurBGswVvhkdvBmG3CuGXx1Z+1AP1iOEGRq8WYfdM8MvLz1IZOq6qRHBW2XYPXOS5FKxD/VgTeq6md4v0Kg47J7d2x77mdUO9dlhH3ekppvoGbyahtV268Hd9ncHLqo20Ct4NQor7Tysdt73ytS0sB7Bu2OD+Cwmjh377yeOQz1qWtTV4NUYrKT/vnM86lHTgq58q3lWQ3y6utSwtUWq763F2Z57c0emGjUtJtKPJBx6NdhP/06P4egEWG92yL1T89/+PzO3ecdBPXN/V71nvmJNK+7Trz4N3t6NN+vgP/6dKCsUzplRtxH9+Pxvjt6PqoM6glfH1UzhLatXvCsHwychbJCtNXqwzazv8bdG7pPwnWv0hb16FvHJl6t6NVSkBrq1WNvDn0YF1K2trf/ov20VtsaouqpnAb3fXPWOyCH3t22Lus07GBm6UYw8LwzrdSLPOxY5G7y7nMBOljhG9Fzk+lop1dSzpmqZ3OwVb9Rh90q27a1mxJf4stQ0y3byPjWltTY3eLM2XaZhzWsZa9i796yU1utVT7VMbMVnvPCOnoMlY+g+Er61ZO9HLjoTvFdOWI3GGUL3T1X2gy/quTErXiqrNtzcpqzlaj3VMSnBSzQVnxPvSfhCcoKXiqqG7qH6/u1ELTc0I3g1Fu/yIxZzWfXCAjOC18kN/bnAoDXzNSW3monCave8HvtrcK+3W99uT/BSheEFpCB4icCq63MuOCCZWcFrsDLS7uHjeVBI5Ezw7j7cAOCymbeaXVXzk6t94YLwi+MAScz+jFf4QkzOTZhkxZernOD0YpUHpHM2eHu+J1YAQ18uRCCB1Y8TCeC9qT2wndXBexDAnGV1159zECb4JHhHDrx7E8JwhQsSCO7TFe+Mk1sIA1DOP6s34E3P4euqHqjiyuIi+yzMvLD6+NhfCd5bW3fQBPHe1BtI6+qKd2X4PhLEQEYR5ieT9fhWc8SQuzefEUenLuNEPCeBf/X6jDfKyvcVK2IgGu8p31TPL1cdTRA5gA+P26h5AZhmxAs0sgWZ29HAbGbOxka9uerWBDDAT3rMmWzzlQejXxkpgAH+Y7Yw7QUamT7/PRzbmu3CAYin5+wzk5Kb/SMJt5ZvFZzpYgGIxwzhm5WvjHwM3+iNafULnDVirplBBUR5V/NzM0UN4nvT+L1Ef/Y7M8d1nZHH3uwpIkrwPoscxMIXOESaTRltOUujBu+zaEEsfNdy/Hm2eiaMpt8LyRK8z6IFMcAoQreYrMH7bEUQW3UBo5kxBc1+nGiWWY8tWWkTjZ6sQ+gWVTV4H2V7bpj3CJj+nCdxqEVhOwTvQQDHox7wnTm1gZ2C99C7sa28iEIv5iZwN7Fj8B40eX6Chgqscjezc/C2ptnh4FyYT+Buavfgba1P41t5fe7q8XfsvzgOeQjczQneL04CAKYQvP8Rvnntvtq7uv96f67d+3V7gpcIDP7PGeLz9bhVrG4bq/LKSPAKz71lrL2e3ZQV73dOgnV8ye28Hvur5z/n2PERwQs57XaREZVbzpwmeInEqncuK7YY9OxmBC8VVR9k1fcvGxcwnCJ4vzPQ1us1xKrWstd+CYu+3HLmbZ8GryYhg2p9KnRrq9avvPBJ8N4f/lejMIJfj/pTlf2ozHcUeEuPW81VGuXKflhB9Cd8/9Nz+/XqWI4vvzobvK8GgNUv0WXsz97nlVCYw+e9/FXvL1dlbZis211d76DIdIGYZTsZQ/0LOxO87zZCpuHWmhfMRzfi+Ebu0VHbpk/ncrx5aeTjRJGH2yH69vFl1BCLVP+R54sQWMMtZ3707o8kXGmAx/82ygDwWEY+tzZmEK3uz9HDVY/m5ocUCpr9Ao3Vq+DVf59rRg+gmf0x428Z2OupAX94Z8U7epXR2pyB2psTag1RK99HI/pz9gWf/ozjas9a9RYT5fd4f2rKTxvNira+GeH7KFtPGdL1VA3fbOfWGS/rFSV4fxK5IBVPgGxmh28WejOmHv1aNXy3885nvAr9neMRh1p853jEpj601t7/cpWG+eI4xKMmXxyHHDxixKlvNe9+Yu++/5Hd2r712XnfdyV8kzv7ONGuJ/mO+5zRbnXabX+rULfNffoc704BvMt+VrFDb+6wj9W55byxq99qPpqnYhMYbLlV7E09ySPfck6q15urql2BV9qX3VXozQr7wJ8yvpiFDno/x5t9lWG41ZWxN/VjfZ5H39CoF2g8DozoTWW47SVDAOtJznDLOZkZb656bogIA0+TEu3iUE/uy1utNnO735fPnBkboCE5a2Rf6kfYWIR3Nb8zhH4bggYZvf3WU3/rSf0IvBQheN9hkBGNngQ+0utxIgDgDYIXACYSvAAwkeAFgIkELwBMJHgBYKL/A12yUoTvAlGCAAAAAElFTkSuQmCC"

CAT_BADGE = {
    'NMN': 'cat-nmn',
    '데일리 릴리프': 'cat-daily',
    'PDRN': 'cat-pdrn',
    '사은품': 'cat-gift',
    '쇼핑백/리플렛': 'cat-pack',
    'GWP': 'cat-gwp',
    '부자재': 'cat-mat',
}
CAT_BADGE_AMZ = {'NMN': 'cat-nmn', '데일리 릴리프': 'cat-daily'}
CH_NAMES = ['약국영업','파라다이스','명동친한약국','카페24자사몰','임직원구매','B2B','마케팅','마케팅시딩']
WEEKDAY_NAMES_KO = ['일', '월', '화', '수', '목', '금', '토']


def fmt_date(v):
    """Dot-separated date, used for legacy display (기준일 / 업데이트 시간)."""
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y.%m.%d')
    return v


def fmt_date_iso(v):
    """Hyphen-separated date (YYYY-MM-DD), used for 입고예정/유통기한/캘린더."""
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    if v is None:
        return None
    return str(v)


def n(v):
    if v is None:
        return '—'
    if isinstance(v, str):
        return v
    if isinstance(v, (int, float)):
        if float(v).is_integer():
            return f'{int(v):,}'
        return f'{v:,.2f}'
    return str(v)


def esc(v):
    if v is None:
        return ''
    return str(v)


def dday_class(dday):
    """Urgency thresholds for 입고 D-Day (shipment arrival)."""
    if dday is None or not isinstance(dday, (int, float)):
        return 'dday-na'
    if dday <= 15:
        return 'dday-urgent'
    if dday <= 35:
        return 'dday-soon'
    return 'dday-later'


def exp_dday_class(days):
    """Urgency thresholds for 유통기한 (expiration)."""
    if days is None:
        return 'dday-na'
    if days <= 90:
        return 'dday-urgent'
    if days <= 180:
        return 'dday-soon'
    return 'dday-later'


def _split_batch(s):
    """Split a raw '30,000 ┃ 10,000' style cell into ['30,000', '10,000']."""
    if not s:
        return []
    return [p.strip() for p in str(s).split('┃') if p.strip()]


def format_incoming_lines(qty_raw, date_raw):
    """
    입고예정 수량/일자 셀은 여러 배치가 '┃'로 이어져 있고 날짜는 '2026.10.23' 같은
    텍스트라서, 각 배치를 'YYYY-MM-DD N,000개' 한 줄씩으로 변환한다.
    """
    qtys = _split_batch(qty_raw)
    dates = _split_batch(date_raw)
    if not qtys:
        return []
    lines = []
    for i, q in enumerate(qtys):
        d = dates[i] if i < len(dates) else None
        d_disp = d.replace('.', '-') if d else '날짜 미정'
        q_disp = q if q.endswith('개') else f'{q}개'
        lines.append(f'{d_disp} {q_disp}')
    return lines


def est_depletion(avail, monthly_total, today):
    """
    Estimate the date available stock runs out.
    monthly_total is a monthly average outbound quantity (computed directly
    from the '출고 Raw' sheet by extract_shipment_raw), so daily rate =
    monthly_total / 30.
    Returns a dict describing the result, or None if avail isn't a number.
    """
    if not isinstance(avail, (int, float)):
        return None
    if avail <= 0:
        return {'status': 'out'}
    if not isinstance(monthly_total, (int, float)) or monthly_total <= 0:
        return {'status': 'no-data'}
    daily_rate = monthly_total / 30.0
    days_left = avail / daily_rate
    est_date = today + datetime.timedelta(days=days_left)
    return {'status': 'ok', 'days': int(days_left), 'date': est_date}


def extract_shipment_raw(wb):
    """
    Aggregate a monthly-average outbound quantity per product code straight
    from the '출고 Raw' sheet, so 예상소진일 no longer depends on the
    pre-computed '채널별 평균 출고량' columns in '★ 국내'.

    '출고 Raw' is laid out as repeated blocks, one per sales channel, each
    starting with a '▣ <채널명>' marker in column A, followed by a header
    row ('제품코드', '품목명', then one column per month such as '4월'),
    followed by one data row per product until the next block. This walks
    every block, sums each product code's quantity across all month columns
    it finds (so it keeps working as new months are appended), and divides
    by the number of distinct months seen to get a monthly average — summed
    across every channel block, matching how the source spreadsheet's own
    per-channel averages are derived.
    """
    ws = wb['출고 Raw']
    totals = {}
    month_labels = set()
    max_row = ws.max_row
    max_col = ws.max_column
    r = 1
    while r <= max_row:
        marker = ws.cell(row=r, column=1).value
        if isinstance(marker, str) and marker.strip().startswith('▣'):
            header_row = r + 1
            headers = [ws.cell(row=header_row, column=c).value for c in range(1, max_col + 1)]
            month_cols = [c for c, h in enumerate(headers, start=1) if isinstance(h, str) and h.strip().endswith('월')]
            month_labels.update(headers[c - 1] for c in month_cols)
            data_r = header_row + 1
            while data_r <= max_row:
                code = ws.cell(row=data_r, column=1).value
                name = ws.cell(row=data_r, column=2).value
                if code is None and name is None:
                    break
                if isinstance(code, (int, float)):
                    code_key = int(round(code))
                    total = totals.get(code_key, 0.0)
                    for c in month_cols:
                        val = ws.cell(row=data_r, column=c).value
                        if isinstance(val, (int, float)):
                            total += val
                    totals[code_key] = total
                data_r += 1
            r = data_r
            continue
        r += 1
    n_months = len(month_labels) or 1
    return {code: total / n_months for code, total in totals.items()}


def download_workbook():
    resp = requests.get(EXPORT_URL, timeout=60, allow_redirects=True)
    resp.raise_for_status()
    ctype = resp.headers.get('Content-Type', '')
    if 'html' in ctype.lower() or resp.content[:5] == b'<!DOC':
        raise RuntimeError(
            "Google Sheet export returned an HTML (login) page instead of XLSX. "
            "The sheet's sharing setting was probably changed back to private. "
            "It needs to stay set to 'Anyone with the link - Viewer'."
        )
    with open(XLSX_PATH, 'wb') as f:
        f.write(resp.content)


def extract_domestic(wb):
    ws = wb['★ 국내']
    updated = ws.cell(row=2, column=3).value
    ref_date_raw = ws.cell(row=3, column=14).value
    rows = []
    for r in range(6, 45):
        name = ws.cell(row=r, column=6).value
        if not name:
            continue
        rows.append({
            'cat': ws.cell(row=r, column=2).value,
            'code': ws.cell(row=r, column=5).value,
            'name': name,
            'lot': ws.cell(row=r, column=7).value,
            'exp_raw': ws.cell(row=r, column=8).value,
            'total': ws.cell(row=r, column=9).value,
            'avail': ws.cell(row=r, column=10).value,
            'blocked': ws.cell(row=r, column=11).value,
            'incoming_qty': ws.cell(row=r, column=12).value,
            'incoming_date': ws.cell(row=r, column=13).value,
            'dday': ws.cell(row=r, column=14).value,
            'note': ws.cell(row=r, column=15).value,
            'ch': {
                '약국영업': ws.cell(row=r, column=16).value,
                '파라다이스': ws.cell(row=r, column=17).value,
                '명동친한약국': ws.cell(row=r, column=18).value,
                '카페24자사몰': ws.cell(row=r, column=19).value,
                '임직원구매': ws.cell(row=r, column=20).value,
                'B2B': ws.cell(row=r, column=21).value,
                '마케팅': ws.cell(row=r, column=22).value,
                '마케팅시딩': ws.cell(row=r, column=23).value,
                '합계': ws.cell(row=r, column=24).value,
            }
        })
    return {
        'updated': str(updated),
        'ref_date': fmt_date(ref_date_raw),
        'ref_date_raw': ref_date_raw,
        'rows': rows,
    }


def extract_amazon(wb):
    ws = wb['★ 아마존']
    updated = ws.cell(row=2, column=3).value
    inv_rows = []
    for r in range(6, 15):
        name = ws.cell(row=r, column=6).value
        if not name:
            continue
        inv_rows.append({
            'cat': ws.cell(row=r, column=2).value,
            'sku': ws.cell(row=r, column=3).value,
            'asin': ws.cell(row=r, column=5).value,
            'name': name,
            'name_amz': ws.cell(row=r, column=7).value,
            'avail': ws.cell(row=r, column=8).value,
            'units_per_box': ws.cell(row=r, column=9).value,
            'box_l': ws.cell(row=r, column=10).value,
            'box_w': ws.cell(row=r, column=11).value,
            'box_h': ws.cell(row=r, column=12).value,
        })
    ship_rows = []
    for r in range(18, 27):
        name = ws.cell(row=r, column=6).value
        if not name:
            continue
        ship_rows.append({'sku': ws.cell(row=r, column=3).value, 'total_shipped': ws.cell(row=r, column=8).value})
    return {'updated': str(updated), 'inv_rows': inv_rows, 'ship_rows': ship_rows}


def _today_ref(domestic):
    ref = domestic.get('ref_date_raw')
    if isinstance(ref, datetime.datetime):
        return ref
    return datetime.datetime.utcnow()


def build_domestic_html(domestic, shipment_avg):
    d_rows = domestic['rows']
    today_ref = _today_ref(domestic)
    total_sku = len(d_rows)
    total_stock = sum((r['total'] or 0) for r in d_rows if isinstance(r['total'], (int, float)))
    blocked_items = [r for r in d_rows if isinstance(r['blocked'], (int, float)) and r['blocked'] > 0]
    total_blocked = sum(r['blocked'] for r in blocked_items)
    incoming_items = [r for r in d_rows if r['incoming_qty']]

    def dday_key(r):
        dd = r['dday']
        return dd if isinstance(dd, (int, float)) else 9999

    incoming_sorted = sorted(incoming_items, key=dday_key)
    dday_errors = [r for r in d_rows if isinstance(r['dday'], str) and 'VALUE' in r['dday']]

    # depletion estimates, computed once so both the table and findings can use them.
    # Monthly outbound rate comes straight from '출고 Raw' (see extract_shipment_raw),
    # matched by 제품코드, not from the '★ 국내' sheet's pre-computed averages.
    depletions = {}
    for r in d_rows:
        code_key = int(round(r['code'])) if isinstance(r['code'], (int, float)) else None
        monthly_avg = shipment_avg.get(code_key) if code_key is not None else None
        depletions[id(r)] = est_depletion(r['avail'], monthly_avg, today_ref)
    urgent_depletions = [
        r for r in d_rows
        if depletions[id(r)] and depletions[id(r)]['status'] == 'ok' and depletions[id(r)]['days'] <= 14
    ]
    urgent_depletions.sort(key=lambda r: depletions[id(r)]['days'])

    groups = OrderedDict()
    for r in d_rows:
        groups.setdefault(r['cat'], []).append(r)

    main_rows = []
    for cat, items in groups.items():
        badge_cls = CAT_BADGE.get(cat, 'cat-mat')
        for i, r in enumerate(items):
            cat_cell = f'<span class="cat-badge {badge_cls}">{esc(cat)}</span>' if i == 0 else ''
            incoming_lines = format_incoming_lines(r['incoming_qty'], r['incoming_date'])
            if incoming_lines:
                incoming = '<br>'.join(esc(l) for l in incoming_lines)
            else:
                incoming = '<span class="val-empty">—</span>'
            dd = r['dday']
            if isinstance(dd, str) and 'VALUE' in dd:
                dday_html = '<span class="dday-chip dday-err">확인필요</span>'
            elif isinstance(dd, (int, float)):
                dday_html = f'<span class="dday-chip {dday_class(dd)}">D-{int(dd)}</span>'
            else:
                dday_html = '<span class="val-empty">—</span>'
            blocked_html = f"<span class='stock-blocked'>{n(r['blocked'])}</span>" if (isinstance(r['blocked'], (int, float)) and r['blocked'] > 0) else '<span class="val-empty">0</span>'
            code_disp = esc(int(r['code'])) if isinstance(r['code'], (int, float)) else esc(r['code'])
            note_disp = esc(r['note']) if r['note'] else '<span class="val-empty">—</span>'

            dep = depletions[id(r)]
            if dep is None:
                deplete_html = '<span class="val-empty">—</span>'
            elif dep['status'] == 'out':
                deplete_html = '<span class="dday-chip dday-err">재고소진</span>'
            elif dep['status'] == 'no-data':
                deplete_html = '<span class="val-empty" title="최근 출고 데이터 없음">출고없음</span>'
            else:
                days = dep['days']
                cls = 'dday-urgent' if days <= 14 else ('dday-soon' if days <= 30 else 'dday-later')
                deplete_html = f"<div>{fmt_date_iso(dep['date'])}</div><div class='sub-date {cls}-text'>약 {days}일 후</div>"

            row_cls = ' class="cat-group-end"' if i == len(items) - 1 else ''
            main_rows.append(f"""<tr{row_cls}>
<td class="cat-col">{cat_cell}</td>
<td class="code-col">{code_disp}</td>
<td class="name-col">{esc(r['name'])}</td>
<td class="num">{n(r['total'])}</td>
<td class="num stock-avail">{n(r['avail'])}</td>
<td class="num">{blocked_html}</td>
<td class="incoming-col">{incoming}</td>
<td class="dday-col">{dday_html}</td>
<td class="deplete-col">{deplete_html}</td>
<td class="note-col">{note_disp}</td>
</tr>""")
    domestic_table = '\n'.join(main_rows)

    ch_rows = [r for r in d_rows if isinstance(r['ch']['합계'], (int, float)) and r['ch']['합계'] > 0]
    ch_table_rows = []
    for r in ch_rows:
        cells = ''.join(f"<td class='num'>{n(r['ch'][c])}</td>" for c in CH_NAMES)
        ch_table_rows.append(f"<tr><td class='col-fix name-fix'>{esc(r['name'])}</td>{cells}<td class='num val-total'>{n(r['ch']['합계'])}</td></tr>")
    channel_table = '\n'.join(ch_table_rows)

    findings = []
    findings.append(f"<div class='finding-item'><div class='finding-icon icon-ok'>✓</div><div class='finding-text'>전체 <strong>{total_sku}개 SKU</strong>, 총재고 합계 <strong>{n(total_stock)}개</strong>.</div></div>")
    if blocked_items:
        names = ', '.join(r['name'].split('] ')[-1] for r in blocked_items[:4])
        findings.append(f"<div class='finding-item'><div class='finding-icon icon-warn'>◐</div><div class='finding-text'><strong>불용(대기) 재고 총 {n(total_blocked)}개</strong> — {len(blocked_items)}개 품목에 집중 ({names}{'...' if len(blocked_items) > 4 else ''}).</div></div>")
    if incoming_sorted:
        nearest = incoming_sorted[0]
        nearest_lines = format_incoming_lines(nearest['incoming_qty'], nearest['incoming_date'])
        nearest_disp = esc(nearest_lines[0]) if nearest_lines else '—'
        findings.append(f"<div class='finding-item'><div class='finding-icon icon-warn'>◐</div><div class='finding-text'>입고 예정 <strong>{len(incoming_items)}건</strong> 중 가장 임박: <strong>{esc(nearest['name'])}</strong> — {nearest_disp} (D-{int(nearest['dday']) if isinstance(nearest['dday'], (int, float)) else '?'})</div></div>")
    if urgent_depletions:
        names = ', '.join(r['name'].split('] ')[-1] for r in urgent_depletions[:4])
        findings.append(f"<div class='finding-item'><div class='finding-icon icon-warn'>◐</div><div class='finding-text'>⚠ <strong>2주 내 재고 소진 예상 {len(urgent_depletions)}개 품목</strong> — {esc(names)}{'...' if len(urgent_depletions) > 4 else ''} (출고 추세 기준 추정치입니다).</div></div>")
    if dday_errors:
        names = ', '.join(r['name'] for r in dday_errors)
        findings.append(f"<div class='finding-item'><div class='finding-icon icon-warn'>◐</div><div class='finding-text'>⚠ <strong>입고예정일 미기재로 D-Day 계산 오류</strong>: {esc(names)} — 원본 시트 확인 필요.</div></div>")
    findings_html = '\n'.join(findings)

    return domestic_table, channel_table, findings_html


def build_amazon_html(amazon):
    inv_rows = amazon['inv_rows']
    ship_map = {r['sku']: r['total_shipped'] for r in amazon['ship_rows']}
    total_avail = sum((r['avail'] or 0) for r in inv_rows)
    sorted_by_avail = sorted(inv_rows, key=lambda r: r['avail'] if isinstance(r['avail'], (int, float)) else 0)
    lowest, highest = sorted_by_avail[0], sorted_by_avail[-1]

    rows = []
    for r in inv_rows:
        badge = CAT_BADGE_AMZ.get(r['cat'], 'cat-mat')
        name_en = r['name_amz'].split('|')[0].strip() if r['name_amz'] else r['name']
        shipped = ship_map.get(r['sku'])
        box_spec = f"{n(r['box_l'])}×{n(r['box_w'])}×{n(r['box_h'])} in · {n(r['units_per_box'])}개/box"
        rows.append(f"""<tr>
<td class="cat-col"><span class="cat-badge {badge}">{esc(r['cat'])}</span></td>
<td class="name-col">{esc(name_en)}<div class='sub-code'>{esc(r['name'])}</div></td>
<td class="code-col mono">{esc(r['sku'])}</td>
<td class="code-col mono">{esc(r['asin'])}</td>
<td class="num stock-avail">{n(r['avail'])}</td>
<td class="num">{n(shipped)}</td>
<td class="box-col">{box_spec}</td>
</tr>""")
    amazon_table = '\n'.join(rows)

    findings = []
    findings.append(f"<div class='finding-item'><div class='finding-icon icon-ok'>✓</div><div class='finding-text'>Amazon US 핵심 SKU <strong>{len(inv_rows)}개</strong>, 가용재고 합계 <strong>{n(total_avail)}개</strong>.</div></div>")
    findings.append(f"<div class='finding-item'><div class='finding-icon icon-warn'>◐</div><div class='finding-text'>가용재고 최다: <strong>{esc(highest['name_amz'].split('|')[0].strip())}</strong> ({n(highest['avail'])}개) · 최소: <strong>{esc(lowest['name_amz'].split('|')[0].strip())}</strong> ({n(lowest['avail'])}개)</div></div>")
    findings_html = '\n'.join(findings)

    return amazon_table, findings_html


def build_expiry_html(domestic):
    d_rows = domestic['rows']
    items = [r for r in d_rows if isinstance(r.get('exp_raw'), datetime.datetime)]
    items.sort(key=lambda r: r['exp_raw'])
    today_ref = _today_ref(domestic).date()

    if not items:
        return '<div class="findings"><p style="color:#9ca3af;">유통기한 정보가 있는 품목이 없습니다.</p></div>', ''

    urgent = 0
    body_rows = []
    for r in items:
        days = (r['exp_raw'].date() - today_ref).days
        cls = exp_dday_class(days)
        if cls == 'dday-urgent':
            urgent += 1
        code_disp = esc(int(r['code'])) if isinstance(r['code'], (int, float)) else esc(r['code'])
        dday_text = f"D-{days}" if days >= 0 else f"만료 {-days}일 경과"
        body_rows.append(f"""<tr>
<td class="code-col">{code_disp}</td>
<td class="name-col">{esc(r['name'])}</td>
<td>{esc(r['lot']) if r['lot'] else '—'}</td>
<td class="exp-col">{fmt_date_iso(r['exp_raw'])}</td>
<td class="num">{n(r['total'])}</td>
<td class="num stock-avail">{n(r['avail'])}</td>
<td class="dday-col"><span class="dday-chip {cls}">{dday_text}</span></td>
</tr>""")
    table = '\n'.join(body_rows)

    findings = [f"<div class='finding-item'><div class='finding-icon icon-ok'>✓</div><div class='finding-text'>유통기한 등록 품목 <strong>{len(items)}개</strong> 중 임박(90일 이내) <strong>{urgent}개</strong>.</div></div>"]
    findings_html = '\n'.join(findings)

    panel = f"""<div class="panel">
<div class="panel-header"><h2>⏳ 유통기한 임박 현황</h2><div class="meta">유통기한 빠른 순 정렬 · 총 {len(items)}개 품목</div></div>
<table>
<thead><tr><th>제품코드</th><th style="text-align:left;">제품명</th><th>LOT</th><th>유통기한</th><th>총재고</th><th>가용재고</th><th>D-Day</th></tr></thead>
<tbody>
{table}
</tbody>
</table>
</div>"""
    return panel, findings_html


def _parse_dot_date(s):
    """Parse a '2026.10.23' text date into a date object, or None if unparseable."""
    if not s:
        return None
    try:
        y, m, d = [int(p) for p in s.strip().split('.')]
        return datetime.date(y, m, d)
    except (ValueError, TypeError):
        return None


def build_calendar_html(domestic):
    d_rows = domestic['rows']
    events = []
    for r in d_rows:
        qtys = _split_batch(r.get('incoming_qty'))
        dates = _split_batch(r.get('incoming_date'))
        for i, q in enumerate(qtys):
            d_str = dates[i] if i < len(dates) else None
            d = _parse_dot_date(d_str)
            if d is not None:
                events.append((d, r['name'], q))

    if not events:
        return '<div class="findings"><p style="color:#9ca3af;">입고 예정 일정이 없습니다.</p></div>'

    events.sort(key=lambda e: e[0])
    months = OrderedDict()
    for ev in events:
        key = (ev[0].year, ev[0].month)
        months.setdefault(key, []).append(ev)

    cal = calmod.Calendar(firstweekday=6)  # weeks start on Sunday
    header_cells = ''.join(f"<div class='cal-dow'>{wd}</div>" for wd in WEEKDAY_NAMES_KO)

    blocks = []
    for (y, m), evs in months.items():
        day_events = {}
        for (date_, name, qty) in evs:
            day_events.setdefault(date_.day, []).append((name, qty))

        weeks = cal.monthdayscalendar(y, m)
        week_html = []
        for week in weeks:
            cells = []
            for day in week:
                if day == 0:
                    cells.append("<div class='cal-cell cal-empty'></div>")
                    continue
                evs_today = day_events.get(day, [])
                ev_html = ''.join(
                    f"<div class='cal-event'>{esc(name)}<span>{n(qty)}개</span></div>"
                    for (name, qty) in evs_today
                )
                today_cls = ' cal-has-event' if evs_today else ''
                cells.append(f"<div class='cal-cell{today_cls}'><div class='cal-daynum'>{day}</div>{ev_html}</div>")
            week_html.append(f"<div class='cal-week'>{''.join(cells)}</div>")

        blocks.append(f"""<div class="panel">
<div class="panel-header"><h2>{y}년 {m}월 입고 일정</h2><div class="meta">{len(evs)}건</div></div>
<div class="cal-grid">
<div class="cal-week cal-dow-row">{header_cells}</div>
{''.join(week_html)}
</div>
</div>""")

    return '\n'.join(blocks)


def assemble(domestic, amazon, domestic_table, channel_table, findings_domestic,
             amazon_table, findings_amazon, expiry_panel, findings_expiry, calendar_html):
    domestic_updated_disp = domestic['updated'][:16].replace('T', ' ')
    amazon_updated_disp = amazon['updated'][:16].replace('T', ' ')
    build_time = datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8"/>
<meta content="width=device-width, initial-scale=1.0" name="viewport"/>
<title>재고 운영 현황 | SIORE</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Apple SD Gothic Neo', 'Noto Sans KR', sans-serif; background: #f4f5f7; color: #1a1a2e; padding: 24px 24px 32px; font-size: 13px; }}
  .page-header {{ background: linear-gradient(135deg, #0d2b45 0%, #1b4d6e 100%); margin: -24px -24px 0; padding: 16px 28px; display: flex; align-items: center; gap: 20px; }}
  .page-header-logo {{ height: 26px; flex-shrink: 0; display: flex; align-items: center; }}
  .page-header-logo img {{ height: 26px; width: auto; }}
  .page-header-sep {{ width: 1px; height: 32px; background: rgba(255,255,255,0.25); flex-shrink: 0; }}
  .page-header-text h1 {{ font-size: 16px; font-weight: 700; color: #ffffff; margin-bottom: 3px; }}
  .page-header-text .subtitle {{ color: rgba(255,255,255,0.55); font-size: 11px; }}
  .tab-nav {{ display: flex; gap: 0; background: #fff; margin: 0 -24px; padding: 0 16px; border-bottom: 1px solid #e5e7eb; box-shadow: 0 2px 8px rgba(0,0,0,0.08); position: relative; flex-wrap: wrap; }}
  .tab-nav::before {{ content: ''; position: absolute; top: 0; left: 0; right: 0; height: 3px; background: linear-gradient(90deg, #14b8a6, #0ea5e9); }}
  .tab-btn {{ padding: 14px 24px 13px; border: none; background: none; cursor: pointer; font-size: 14px; font-weight: 600; color: #9ca3af; border-bottom: 3px solid transparent; margin-bottom: -1px; transition: color 0.18s, border-color 0.18s; font-family: inherit; white-space: nowrap; letter-spacing: -0.01em; }}
  .tab-btn:hover {{ color: #374151; border-bottom-color: #d1fae5; }}
  .tab-btn.active {{ color: #0d9488; font-weight: 700; border-bottom-color: #14b8a6; }}
  .tab-content {{ display: none; }}
  .tab-content.active {{ display: block; }}
  .tab-panel-wrap {{ padding-top: 20px; }}
  .findings {{ background: #fff; border-radius: 12px; padding: 20px 24px; margin-bottom: 16px; box-shadow: 0 1px 4px rgba(0,0,0,0.07); }}
  .findings h2 {{ font-size: 14px; font-weight: 700; margin-bottom: 14px; }}
  .finding-item {{ display: flex; gap: 12px; margin-bottom: 10px; align-items: flex-start; }}
  .finding-icon {{ width: 22px; height: 22px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 12px; flex-shrink: 0; margin-top: 1px; }}
  .icon-ok {{ background: #dcfce7; }}
  .icon-warn {{ background: #fef3c7; }}
  .finding-text {{ font-size: 12.5px; color: #374151; line-height: 1.55; }}
  .finding-text strong {{ color: #111827; }}
  .panel-header {{ display:flex; align-items:center; justify-content:space-between; gap:12px; padding:14px 20px; border-bottom:1px solid #f0f0f4; flex-wrap:wrap; }}
  .panel-header h2 {{ font-size: 14px; font-weight: 700; }}
  .panel-header .meta {{ font-size: 11px; color: #9ca3af; }}
  .panel {{ background:#fff; border-radius:12px; margin-bottom:16px; box-shadow:0 1px 4px rgba(0,0,0,0.07); overflow:hidden; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 12.5px; }}
  thead th {{ background: #f9fafb; padding: 10px 12px; text-align: center; font-weight: 600; color: #6b7280; font-size: 11px; letter-spacing: 0.03em; border-bottom: 1px solid #e5e7eb; white-space: nowrap; }}
  thead th:first-child {{ text-align: left; }}
  th .th-sub {{ display: block; font-weight: 400; color: #b0b6c0; font-size: 9.5px; letter-spacing: 0; white-space: normal; line-height: 1.35; margin-top: 3px; }}
  tbody tr:hover {{ background: #fafafa; }}
  tbody td {{ padding: 9px 12px; border-bottom: 1px solid #f3f4f6; vertical-align: middle; line-height: 1.45; }}
  .stock-table {{ table-layout: fixed; min-width: 1080px; }}
  .stock-table th, .stock-table td {{ overflow-wrap: break-word; white-space: normal; }}
  .stock-table .code-col, .stock-table .dday-col, .stock-table .incoming-col, .stock-table .deplete-col {{ white-space: nowrap; }}
  .cat-group-end > td {{ border-bottom: 2px solid #d8dbe1; }}
  td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
  .name-col {{ text-align:left; font-weight:600; color:#1f2937; min-width:180px; }}
  .code-col {{ font-size:11px; color:#9ca3af; font-family:monospace; text-align:left; }}
  .exp-col {{ font-size:11.5px; color:#6b7280; text-align:left; white-space:nowrap; }}
  .cat-col {{ text-align:left; }}
  .incoming-col {{ font-size:11.5px; color:#6b7280; line-height:1.5; text-align:left; white-space:nowrap; }}
  .incoming-col .sub-date {{ color:#9ca3af; font-size:11px; }}
  .deplete-col {{ font-size:11.5px; color:#6b7280; line-height:1.5; text-align:left; white-space:nowrap; }}
  .deplete-col .sub-date {{ font-size:11px; }}
  .dday-urgent-text {{ color:#dc2626; font-weight:600; }}
  .dday-soon-text {{ color:#d97706; font-weight:600; }}
  .dday-later-text {{ color:#9ca3af; }}
  .dday-col {{ text-align:center; }}
  .box-col {{ font-size:11px; color:#6b7280; text-align:left; white-space:nowrap; }}
  .mono {{ font-family:monospace; }}
  .sub-code {{ font-size:11px; color:#9ca3af; font-weight:400; margin-top:2px; }}
  .note-col {{ font-size:11.5px; color:#c2410c; font-weight:500; text-align:left; line-height:1.4; }}
  .val-empty {{ color:#d1d5db; font-size:11px; }}
  .cat-badge {{ display:inline-block; padding:3px 10px; border-radius:20px; font-size:11px; font-weight:700; white-space:nowrap; }}
  .cat-nmn {{ background:#eff6ff; color:#1d4ed8; }}
  .cat-daily {{ background:#f0fdf4; color:#166534; }}
  .cat-pdrn {{ background:#fdf4ff; color:#a21caf; }}
  .cat-gift {{ background:#fdf2f8; color:#be185d; }}
  .cat-pack {{ background:#fff7ed; color:#c2410c; }}
  .cat-gwp {{ background:#f0fdfa; color:#0d9488; }}
  .cat-mat {{ background:#f3f4f6; color:#4b5563; }}
  .stock-avail {{ font-weight:700; color:#111827; }}
  .stock-blocked {{ color:#dc2626; font-weight:700; }}
  .dday-chip {{ display:inline-block; padding:2px 9px; border-radius:10px; font-size:11px; font-weight:700; }}
  .dday-urgent {{ background:#fee2e2; color:#dc2626; }}
  .dday-soon {{ background:#fef3c7; color:#d97706; }}
  .dday-later {{ background:#f3f4f6; color:#6b7280; }}
  .dday-err {{ background:#fee2e2; color:#dc2626; }}
  .dday-na {{ background:#f3f4f6; color:#9ca3af; }}
  .ch-scroll {{ overflow-x: auto; }}
  .ch-table {{ width: max-content; min-width: 100%; border-collapse: collapse; font-size: 12px; }}
  .ch-table th {{ padding: 8px 12px; text-align: right; font-weight: 600; color: #6b7280; font-size: 11px; background: #f9fafb; border-bottom: 1px solid #e5e7eb; white-space: nowrap; }}
  .ch-table th.col-fix {{ text-align: left; position: sticky; left: 0; background: #f9fafb; z-index: 2; min-width: 200px; }}
  .ch-table td {{ padding: 9px 12px; border-bottom: 1px solid #f3f4f6; text-align: right; white-space: nowrap; font-variant-numeric: tabular-nums; }}
  .ch-table td.col-fix {{ text-align: left; position: sticky; left: 0; background: #fff; font-weight:600; z-index: 1; min-width:200px; }}
  .ch-table tbody tr:hover td {{ background: #fafafa; }}
  .ch-table tbody tr:hover td.col-fix {{ background: #fafafa; }}
  .val-total {{ color: #1d4ed8; font-weight: 700; }}
  .updated-note {{ padding: 12px 20px; font-size: 11.5px; color: #6b7280; border-top: 1px solid #f3f4f6; background: #fafafa; }}
  .build-footer {{ text-align:center; font-size:11px; color:#9ca3af; padding: 18px 0 4px; }}
  .cal-grid {{ padding: 12px 20px 20px; }}
  .cal-week {{ display: grid; grid-template-columns: repeat(7, 1fr); gap: 6px; margin-bottom: 6px; }}
  .cal-dow-row {{ margin-bottom: 10px; }}
  .cal-dow {{ text-align:center; font-size:11px; font-weight:700; color:#9ca3af; padding-bottom:4px; }}
  .cal-cell {{ height: 104px; box-sizing: border-box; overflow-y: auto; border: 1px solid #f0f0f4; border-radius: 8px; padding: 6px 6px; background:#fafafa; }}
  .cal-cell::-webkit-scrollbar {{ width: 4px; }}
  .cal-cell::-webkit-scrollbar-thumb {{ background: #e5e7eb; border-radius: 2px; }}
  .cal-cell.cal-empty {{ background: transparent; border-color: transparent; }}
  .cal-cell.cal-has-event {{ background:#f0fdfa; border-color:#99f6e4; }}
  .cal-daynum {{ font-size: 11px; color:#9ca3af; font-weight:700; margin-bottom:4px; }}
  .cal-event {{ font-size: 10.5px; background:#0d9488; color:#fff; border-radius:6px; padding:3px 6px; margin-bottom:3px; display:flex; justify-content:space-between; gap:6px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
  .cal-event span {{ font-weight:700; flex-shrink:0; }}
</style>
</head>
<body>
<div class="page-header">
  <div class="page-header-logo"><img src="{LOGO_DATAURI}" alt="SIORÉ" /></div>
  <div class="page-header-sep"></div>
  <div class="page-header-text">
    <h1>재고 운영 현황</h1>
    <p class="subtitle">국내 재고 · 아마존(US) 재고 &nbsp;|&nbsp; 기준일: {domestic['ref_date']}</p>
  </div>
</div>
<div class="tab-nav">
<button class="tab-btn active" onclick="switchTab('tab-domestic')">📦 국내 재고 현황</button>
<button class="tab-btn" onclick="switchTab('tab-amazon')">🌎 아마존 재고 현황</button>
<button class="tab-btn" onclick="switchTab('tab-expiry')">⏳ 유통기한 임박 현황</button>
<button class="tab-btn" onclick="switchTab('tab-calendar')">📅 입고 일정</button>
</div>
<div class="tab-content active" id="tab-domestic">
<div class="tab-panel-wrap">
<div class="findings">
<h2>🔍 핵심 요약</h2>
{findings_domestic}
</div>
<div class="panel">
<div class="panel-header">
<h2>📦 제품별 재고 현황</h2>
<div class="meta">재고 업데이트: {domestic_updated_disp}</div>
</div>
<div class="ch-scroll">
<table class="stock-table">
<colgroup>
<col style="width:7%"><col style="width:10%"><col style="width:19%"><col style="width:7%"><col style="width:7%"><col style="width:8%"><col style="width:14%"><col style="width:6%"><col style="width:10%"><col style="width:12%">
</colgroup>
<thead><tr><th>구분</th><th>제품코드</th><th style="text-align:left;">제품명</th><th>총재고</th><th>가용재고</th><th>불용(대기)<span class="th-sub">타 채널 할당<br>완료 수량</span></th><th>입고예정</th><th>D-Day</th><th>예상소진일</th><th style="text-align:left;">비고</th></tr></thead>
<tbody>
{domestic_table}
</tbody>
</table>
</div>
</div>
<div class="panel">
<div class="panel-header">
<h2>📊 채널별 평균 출고량</h2>
<div class="meta">합계 &gt; 0인 품목만 표시 · 약국영업/파라다이스/명동친한약국/카페24자사몰/임직원구매/B2B/마케팅/마케팅시딩</div>
</div>
<div class="ch-scroll">
<table class="ch-table">
<thead><tr><th class="col-fix">제품명</th><th>약국영업</th><th>파라다이스</th><th>명동친한약국</th><th>카페24자사몰</th><th>임직원구매</th><th>B2B</th><th>마케팅</th><th>마케팅시딩</th><th>합계</th></tr></thead>
<tbody>
{channel_table}
</tbody>
</table>
</div>
</div>
</div>
</div>
<div class="tab-content" id="tab-amazon">
<div class="tab-panel-wrap">
<div class="findings">
<h2>🔍 핵심 요약</h2>
{findings_amazon}
</div>
<div class="panel">
<div class="panel-header">
<h2>🌎 Amazon US 재고 현황</h2>
<div class="meta">재고 업데이트: {amazon_updated_disp}</div>
</div>
<div class="ch-scroll">
<table>
<thead><tr><th>구분</th><th style="text-align:left;">제품명</th><th>SKU</th><th>ASIN</th><th>가용재고</th><th>출고 집계</th><th>박스 사양</th></tr></thead>
<tbody>
{amazon_table}
</tbody>
</table>
</div>
</div>
<div class="updated-note">
💡 위 "출고 집계"는 원본 시트의 누적 출고 집계 값이며, 월별 세부 출고 데이터는 원본에 아직 반영되지 않아 이번 리포트에서는 제외했습니다.
</div>
</div>
</div>
<div class="tab-content" id="tab-expiry">
<div class="tab-panel-wrap">
<div class="findings">
<h2>🔍 핵심 요약</h2>
{findings_expiry}
</div>
{expiry_panel}
</div>
</div>
<div class="tab-content" id="tab-calendar">
<div class="tab-panel-wrap">
{calendar_html}
</div>
</div>
<div class="build-footer">자동 생성됨 · 매일 오전 10시(KST) 구글시트에서 갱신 · 마지막 빌드: {build_time}</div>
<script>
function switchTab(id) {{
  document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(el => el.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  event.currentTarget.classList.add('active');
}}
</script>
</body>
</html>
"""
    return html


def main():
    download_workbook()
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    domestic = extract_domestic(wb)
    amazon = extract_amazon(wb)
    shipment_avg = extract_shipment_raw(wb)
    domestic_table, channel_table, findings_domestic = build_domestic_html(domestic, shipment_avg)
    amazon_table, findings_amazon = build_amazon_html(amazon)
    expiry_panel, findings_expiry = build_expiry_html(domestic)
    calendar_html = build_calendar_html(domestic)
    html = assemble(domestic, amazon, domestic_table, channel_table, findings_domestic,
                     amazon_table, findings_amazon, expiry_panel, findings_expiry, calendar_html)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"Wrote {OUTPUT_PATH} ({len(html)} bytes)")


if __name__ == "__main__":
    sys.exit(main())
